import streamlit as st
import openpyxl
import datetime
import pandas as pd
import msoffcrypto
from pathlib import WindowsPath
import tempfile
import xlrd
import sys

st.title('週次計画展開')

df = pd.read_csv('directory.txt',header=None)

df1 = pd.read_csv('directory.txt',header=None,nrows=2)
sel1=st.selectbox('書き換えるファイルを選択していください',df1)
'選択ファイル>>>',sel1

df2 = pd.read_csv('directory.txt',header=None,skiprows=[0,1,4,5])
sel2=st.selectbox('書き換えるシートを選択してください',df2)
'選択ファイル>>>',sel2

sel3 = st.text_input(label='4桁のパスワードを入力してください')

if sel3:
   '入力パスワード>>>',sel3
else: 
   st.markdown(f'パスワードが入力されていません')
   sys.exit()


df3 = pd.read_csv('directory.txt',header=None,skiprows=[2,3,4,5])
sel4=st.selectbox('読み込むファイルを選択してください',df3)
'選択ファイル>>>',sel4

df4 = pd.read_csv('directory.txt',header=None,skiprows=[0,1,4,5])
sel5=st.selectbox('読み込むシートを選択してください',df4)
'選択ファイル>>>',sel5
   
df5 = pd.read_csv('directory.txt',header=None,skiprows=[0,1,2,3,5])
sel6=st.selectbox('対象機種を選択してください',df5)
'選択ファイル>>>',sel6

#'2022-12-12 00:00:00'
input_text = st.text_input(label='ここに日付を入力してください yyyy-mm-dd 00:00:00')
button_click = st.button('実行')

if button_click:

    if input_text:
        # パスワード
        PASSWORD = sel3
        # 暗号化されたファイル
        encrypted_file_name = sel1
        # 復号化して保存するファイル
        decrypted_file_name = sel1+'パスワード解除後.xlsx'
        
        # 暗号化ファイルを開く
        f = open(encrypted_file_name, "rb")
        encrypted_file = msoffcrypto.OfficeFile(f)
        # 復号化のパスワードを設定する
        encrypted_file.load_key(password=PASSWORD)
        
        # 復号化したファイルを別のファイルに保存する
        decrypted_file = open(decrypted_file_name, "wb")
        encrypted_file.decrypt(decrypted_file)
        workbook_0 = openpyxl.load_workbook(decrypted_file_name, data_only=True)
        sheet_0 = workbook_0[sel2]
    
        workbook_1 = openpyxl.load_workbook(sel4, data_only=True)
        sheet_1 = workbook_1[sel5]

        for i in range(116,148):
            cell_value = sheet_1.cell(row=i, column=1).value
            if cell_value==sel6:
             #gyo='行目です'
             #print(str(i)+gyo)
             break

        for j in range(130,183):
            cell_value = sheet_1.cell(row=9, column=j).value

            if cell_value==pd.to_datetime(input_text):

             retsu='列目です'
             #print(str(j)+retsu)
             break

        #st.markdown(f'{input_text}"の判定結果は"{str(j)+retsu}')
        for k in range(24,56):
            cell_value = sheet_0.cell(row=k, column=1).value
            if cell_value==sel6:
            #gyo='行目です'
            #print(str(k)+gyo)
             break

        for l in range(150,188):
            cell_value = sheet_0.cell(row=k-3, column=l).value

            if cell_value==pd.to_datetime(input_text):

            #retsu='列目です'
            #print(str(l)+retsu)
             break

        n=0
        for m in range(j,j+24):
            
            bachelor = sheet_1.cell(row = i, column = m).value
            sheet_0.cell(row = k, column = l+n, value = bachelor)
            n += 1
        workbook_0.save(filename = sel1+'入力後.xlsx'
                                    )
        st.markdown(f'{input_text}"の展開は完了しました！"')

    else:
        st.markdown(f'日付を入力してから実行ボタンを押して下さい')
        sys.exit()